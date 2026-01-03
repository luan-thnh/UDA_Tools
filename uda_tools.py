#!/usr/bin/env python3
"""
UDA Tools Pro - Integrated Tool Suite
- Tab 1: Nhap diem UDA
- Tab 2: HRM Auto Check-in
"""
import customtkinter as ctk
from tkinter import filedialog, messagebox, scrolledtext
import time
import threading
import os
import shutil
import openpyxl
import sys
import json
import logging
import random
import subprocess
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException, TimeoutException

# =====================================================
# ============== CAU HINH LOGGING =====================
# =====================================================
APP_DIR = os.path.join(os.path.expanduser("~"), ".uda_tools")
os.makedirs(APP_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(APP_DIR, 'app.log'), encoding='utf-8'),
    ]
)
logger = logging.getLogger(__name__)

# =====================================================
# ============== CAU HINH HE THONG ====================
# =====================================================
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

CONFIG_FILE = os.path.join(APP_DIR, "config.json")
HISTORY_FILE = os.path.join(APP_DIR, "history.json")
VERSION = "4.0.0"

# UDA Config
DEFAULT_TITLES = ["KTTX", "CCAN", "GHP", "THI1"]
ALL_TITLES = ['CCAN', 'KTTX', 'GHP', 'TDNH', 'THTN', 'TLDA', 'THI1']
EXCEL_MAP = {
    "KTTX": "KTTX", "CCAN": "CCAN", "GHP": "GHP",
    "TDNH": "TDNH", "THTN": "THTN", "TLDA": "TLDA", "THI1": "THI1"
}
NHAP_DIEM_URL = "https://uda.edu.vn/cbgv/gv_nhapdiem"

# HRM Config
HRM_BASE_URL = "https://hrm.donga.edu.vn"
HRM_LOGIN_URL = f"{HRM_BASE_URL}/nhan-vien/dang-nhap"
HRM_TASK_URL = f"{HRM_BASE_URL}/social/home/congviecngay"

# =====================================================
# ============== HELPER FUNCTIONS =====================
# =====================================================
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def detect_chrome_path():
    """Auto detect Chrome/Chromium"""
    candidates = []
    if sys.platform == "win32":
        candidates = [
            os.path.expandvars(r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
            os.path.expandvars(r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe"),
            os.path.expandvars(r"%LocalAppData%\Google\Chrome\Application\chrome.exe"),
            os.path.expandvars(r"%ProgramFiles%\BraveSoftware\Brave-Browser\Application\brave.exe"),
        ]
    elif sys.platform == "darwin":
        candidates = [
            "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
            "/Applications/Chromium.app/Contents/MacOS/Chromium",
            "/Applications/Brave Browser.app/Contents/MacOS/Brave Browser",
        ]
    else:
        candidates = [
            "/usr/bin/google-chrome",
            "/usr/bin/google-chrome-stable",
            "/usr/bin/chromium",
            "/usr/bin/chromium-browser",
            "/usr/bin/brave-browser",
            "/opt/chromium.org/thorium/thorium-browser",
            "/snap/bin/chromium",
        ]
    
    for path in candidates:
        if os.path.exists(path):
            return path
    return None

def load_config():
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"Load config error: {e}")
    return {}

def save_config(config):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"Save config error: {e}")

def load_history():
    try:
        if os.path.exists(HISTORY_FILE):
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return []

def save_history(history):
    try:
        # Keep only last 100 entries
        history = history[-100:]
        with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def add_history(entry_type, status, message):
    history = load_history()
    history.append({
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "type": entry_type,
        "status": status,
        "message": message
    })
    save_history(history)

def get_chrome_driver(chrome_path=None, headless=True):
    """Create Chrome driver with auto-detect"""
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    
    browser_path = chrome_path or detect_chrome_path()
    if browser_path and os.path.exists(browser_path):
        options.binary_location = browser_path
    
    if headless:
        options.add_argument("--headless=new")
    
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(30)
    return driver

# =====================================================
# ============== UDA GRADER LOGIC =====================
# =====================================================
def safe_score(val):
    try:
        if val is None:
            return "0.0"
        score = float(str(val).replace(",", "."))
        score = max(0, min(10, score))
        return "{:.1f}".format(score)
    except (ValueError, TypeError):
        return "0.0"

def read_excel_openpyxl(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active 
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            raise ValueError("File Excel rong!")

        headers = [str(h).strip().upper() if h else "" for h in rows[0]]
        
        if "IDSV" not in headers:
            raise ValueError("Thieu cot 'IDSV'!")

        idsv_index = headers.index("IDSV")
        data_list = []
        
        for row in rows[1:]:
            if len(row) <= idsv_index or row[idsv_index] is None:
                continue
            row_data = {"IDSV": str(row[idsv_index]).strip()}
            for i, cell_val in enumerate(row):
                if i < len(headers) and headers[i] in EXCEL_MAP:
                    row_data[headers[i]] = cell_val
            data_list.append(row_data)
        
        return data_list
    except Exception as e:
        raise ValueError(f"Loi doc Excel: {str(e)}")

def run_uda_grader(username, password, monhoc, excel_file, selected_titles, 
                   status_callback, progress_callback=None, is_delete_mode=False, 
                   headless=False, chrome_path=None):
    """Main UDA grading function"""
    action_name = "XOA" if is_delete_mode else "NHAP"
    driver = None
    
    try:
        status_callback("Doc file Excel...")
        if progress_callback:
            progress_callback(5)
        student_data_list = read_excel_openpyxl(excel_file)
        student_map = {item['IDSV']: item for item in student_data_list}

        status_callback("Khoi dong trinh duyet...")
        if progress_callback:
            progress_callback(10)
        
        driver = get_chrome_driver(chrome_path, headless)
        wait = WebDriverWait(driver, 20)

        # Login
        status_callback("Dang nhap...")
        if progress_callback:
            progress_callback(15)
        driver.get("https://uda.edu.vn/default")
        wait.until(EC.presence_of_element_located((By.NAME, "User"))).send_keys(username)
        driver.find_element(By.NAME, "Password").send_keys(password)
        driver.find_element(By.ID, "Lnew1").click()
        time.sleep(1)

        # Navigate to grade page
        status_callback("Mo trang nhap diem...")
        if progress_callback:
            progress_callback(20)
        driver.get(NHAP_DIEM_URL)

        # Select course
        status_callback("Chon mon hoc...")
        if progress_callback:
            progress_callback(25)
        select = wait.until(EC.presence_of_element_located((By.NAME, "ctl00$MainContent$Dmonlop")))
        select.click()
        time.sleep(0.5)
        
        try:
            driver.find_element(By.XPATH, f'//option[@value="{monhoc}"]').click()
        except Exception:
            raise ValueError(f"Khong tim thay mon: {monhoc}")
        
        time.sleep(0.5)
        driver.find_element(By.ID, "MainContent_Lopen").click()
        time.sleep(1.5)

        # Parse grade table
        status_callback("Phan tich bang diem...")
        if progress_callback:
            progress_callback(30)
        tbody = wait.until(EC.presence_of_element_located((By.TAG_NAME, "tbody")))
        rows = tbody.find_elements(By.TAG_NAME, "tr")

        WEB_INDEX = {}
        IDSV_INDEX = None
        header_row_idx = None

        for r_idx, row in enumerate(rows):
            cells = row.find_elements(By.XPATH, ".//th|.//td")
            if not cells:
                continue
            texts = [c.text.strip().upper() for c in cells]
            for i, t in enumerate(texts):
                if "IDSV" in t or "MSSV" in t:
                    IDSV_INDEX = i
            if IDSV_INDEX is None:
                continue
            for i, t in enumerate(texts):
                for key in selected_titles:
                    if key in t:
                        WEB_INDEX[key] = i
            if WEB_INDEX:
                header_row_idx = r_idx
                break

        if header_row_idx is None:
            raise ValueError("Khong tim thay header bang diem")

        # Process grades
        status_callback(f"Dang {action_name} diem...")
        count = 0
        data_rows = rows[header_row_idx + 1:]
        matched_count = 0

        for i, row in enumerate(data_rows):
            tds = row.find_elements(By.TAG_NAME, "td")
            if len(tds) <= IDSV_INDEX:
                continue
            idsv_web = tds[IDSV_INDEX].text.strip()
            
            if idsv_web not in student_map:
                continue
            
            matched_count += 1
            student_info = student_map[idsv_web]

            for key, idx in WEB_INDEX.items():
                td = tds[idx]
                inputs = td.find_elements(By.TAG_NAME, "input")
                if not inputs:
                    continue
                
                target_value = "" if is_delete_mode else safe_score(student_info.get(EXCEL_MAP[key], 0))
                current_val = inputs[0].get_attribute('value')
                
                if current_val != target_value:
                    inputs[0].clear()
                    if target_value:
                        inputs[0].send_keys(target_value)
            
            count += 1
            
            if progress_callback:
                progress = 30 + int((count / max(matched_count, len(student_map))) * 60)
                progress_callback(min(progress, 90))

        # Save
        status_callback("Luu...")
        if progress_callback:
            progress_callback(95)
        save_btn = wait.until(EC.presence_of_element_located((By.ID, "MainContent_Lsave")))
        driver.execute_script("arguments[0].scrollIntoView(true);", save_btn)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", save_btn)
        time.sleep(1)
        
        if progress_callback:
            progress_callback(100)
        status_callback("Hoan tat!")
        
        add_history("UDA", "SUCCESS", f"{action_name} diem cho {count} SV")
        messagebox.showinfo("Thanh cong", f"Da {action_name.lower()} diem cho {count} sinh vien!")

    except Exception as e:
        status_callback(f"Loi: {str(e)[:50]}")
        add_history("UDA", "ERROR", str(e)[:100])
        messagebox.showerror("Loi", str(e))
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass

# =====================================================
# ============== HRM AUTO LOGIC =======================
# =====================================================
def run_hrm_checkin(username, password, task_title, task_detail, 
                    status_callback, chrome_path=None, headless=True):
    """HRM Auto Check-in function"""
    driver = None
    date_str = datetime.now().strftime('%d/%m/%Y')
    
    try:
        status_callback("Khoi dong trinh duyet...")
        driver = get_chrome_driver(chrome_path, headless)
        wait = WebDriverWait(driver, 15)

        # Login
        status_callback("Dang nhap HRM...")
        driver.get(HRM_LOGIN_URL)
        wait.until(EC.presence_of_element_located((By.ID, "username"))).send_keys(username)
        driver.find_element(By.ID, "password").send_keys(password)
        driver.find_element(By.XPATH, '//*[@id="form"]/button').click()
        
        wait.until(EC.url_contains("/social/home"))
        status_callback("Da dang nhap!")

        # Open task page
        driver.get(HRM_TASK_URL)
        status_callback("Mo trang cong viec...")

        # Open form
        form_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="bscv"]/div/div/div/div[2]/div/button')))
        form_btn.click()
        status_callback("Mo form nhap cong viec...")

        # Fill form
        wait.until(EC.presence_of_element_located((By.ID, "congviec"))).send_keys(task_title)
        
        date_input = driver.find_element(By.ID, "thoigian")
        date_input.clear()
        date_input.send_keys(date_str)
        
        # Fill iframe content
        iframe = wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, "iframe[title='Bộ soạn thảo văn bản có định dạng, baocao']")))
        driver.switch_to.frame(iframe)
        body = wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        body.clear()
        body.send_keys(task_detail)
        driver.switch_to.default_content()
        
        status_callback("Da dien noi dung...")

        # Save
        save_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[value='Lưu']")))
        driver.execute_script("arguments[0].scrollIntoView(true);", save_btn)
        time.sleep(0.5)
        save_btn.click()
        time.sleep(2)
        
        status_callback("Da luu thanh cong!")
        add_history("HRM", "SUCCESS", f"Check-in: {task_title}")
        return True

    except Exception as e:
        status_callback(f"Loi: {str(e)[:50]}")
        add_history("HRM", "ERROR", str(e)[:100])
        logger.error(f"HRM Error: {e}")
        return False
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass

# =====================================================
# ============== SETTINGS WINDOW ======================
# =====================================================
class SettingsWindow(ctk.CTkToplevel):
    def __init__(self, parent, config, on_save_callback):
        super().__init__(parent)
        
        self.cfg = config.copy()
        self.on_save_callback = on_save_callback
        self.detected_chrome = detect_chrome_path()
        
        self.title("⚙️ Cai dat trinh duyet")
        self.geometry("600x320")
        self.resizable(False, False)
        
        self._build_ui()
        
        self.transient(parent)
        self.grab_set()
        self.update()
        x = parent.winfo_x() + (parent.winfo_width() - 600) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 320) // 2
        self.geometry(f"600x320+{x}+{y}")
        self.lift()
        self.focus_force()
    
    def _build_ui(self):
        # Main container
        container = ctk.CTkFrame(self, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=20, pady=15)
        
        # Detection Card
        detect_card = ctk.CTkFrame(container, corner_radius=15)
        detect_card.pack(fill="x", pady=(0, 15))
        
        if self.detected_chrome:
            detect_icon = "✅"
            detect_text = f"Da phat hien: {os.path.basename(self.detected_chrome)}"
            detect_color = ("#22c55e", "#16a34a")
        else:
            detect_icon = "⚠️"
            detect_text = "Khong tim thay Chrome tu dong!"
            detect_color = ("#ef4444", "#f87171")
        
        ctk.CTkLabel(detect_card, text=f"{detect_icon} {detect_text}", 
                     text_color=detect_color,
                     font=ctk.CTkFont(size=13, weight="bold")).pack(padx=20, pady=15)
        
        # Custom Path Card
        path_card = ctk.CTkFrame(container, corner_radius=15)
        path_card.pack(fill="x", pady=10)
        path_card.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(path_card, text="🌐 DUONG DAN TRINH DUYET", 
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=("#4f46e5", "#a5b4fc")).grid(row=0, column=0, columnspan=3, 
                                                            sticky="w", padx=20, pady=(15, 10))
        
        ctk.CTkLabel(path_card, text="Custom Path:", font=ctk.CTkFont(size=12)).grid(row=1, column=0, sticky="w", padx=20, pady=10)
        self.chrome_entry = ctk.CTkEntry(path_card, placeholder_text="De trong = tu dong detect", 
                                          height=38, corner_radius=10)
        self.chrome_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=10)
        ctk.CTkButton(path_card, text="📂", width=45, height=38, corner_radius=10,
                      fg_color=("#6366f1", "#4f46e5"),
                      command=self._browse).grid(row=1, column=2, padx=(5, 20), pady=10)
        
        if self.cfg.get("chrome_path"):
            self.chrome_entry.insert(0, self.cfg["chrome_path"])
        
        # Help text
        ctk.CTkLabel(path_card, text="💡 De trong neu da phat hien Chrome. Chi nhap khi dung browser khac.", 
                     font=ctk.CTkFont(size=11),
                     text_color=("#6b7280", "#9ca3af")).grid(row=2, column=0, columnspan=3, 
                                                             sticky="w", padx=20, pady=(0, 15))
        
        # Buttons
        btn_frame = ctk.CTkFrame(container, fg_color="transparent")
        btn_frame.pack(fill="x", pady=10)
        
        ctk.CTkButton(btn_frame, text="💾 Luu", width=120, height=40, corner_radius=10,
                      fg_color=("#22c55e", "#16a34a"), hover_color=("#15803d", "#14532d"),
                      font=ctk.CTkFont(weight="bold"),
                      command=self._save).pack(side="right", padx=5)
        ctk.CTkButton(btn_frame, text="Huy", width=100, height=40, corner_radius=10,
                      fg_color=("gray60", "gray40"),
                      command=self.destroy).pack(side="right", padx=5)
    
    def _browse(self):
        ft = [("Executable", "*.exe"), ("All", "*.*")] if sys.platform == "win32" else [("All", "*")]
        f = filedialog.askopenfilename(title="Chon Chrome/Browser", filetypes=ft)
        if f:
            self.chrome_entry.delete(0, "end")
            self.chrome_entry.insert(0, f)
    
    def _save(self):
        cp = self.chrome_entry.get().strip()
        if cp and not os.path.exists(cp):
            messagebox.showerror("Loi", f"Path khong ton tai:\n{cp}")
            return
        
        if cp:
            self.cfg["chrome_path"] = cp
        else:
            self.cfg.pop("chrome_path", None)
        
        self.on_save_callback(self.cfg)
        messagebox.showinfo("Thanh cong", "Da luu cai dat!")
        self.destroy()

# =====================================================
# ============== HISTORY WINDOW =======================
# =====================================================
class HistoryWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("📋 Lich su hoat dong")
        self.geometry("750x550")
        
        self._build_ui()
        
        self.transient(parent)
        self.update()
        x = parent.winfo_x() + (parent.winfo_width() - 750) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 550) // 2
        self.geometry(f"750x550+{x}+{y}")
        self.lift()
        self.focus_force()
    
    def _build_ui(self):
        # Main container
        container = ctk.CTkFrame(self, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Header Card
        header_card = ctk.CTkFrame(container, corner_radius=15)
        header_card.pack(fill="x", pady=(0, 15))
        
        header_inner = ctk.CTkFrame(header_card, fg_color="transparent")
        header_inner.pack(fill="x", padx=20, pady=15)
        
        ctk.CTkLabel(header_inner, text="📜 LICH SU HOAT DONG", 
                     font=ctk.CTkFont(size=16, weight="bold"),
                     text_color=("#4f46e5", "#a5b4fc")).pack(side="left")
        
        ctk.CTkButton(header_inner, text="🗑️ Xoa lich su", width=120, height=35, corner_radius=10,
                      fg_color=("#ef4444", "#dc2626"), hover_color=("#b91c1c", "#991b1b"),
                      font=ctk.CTkFont(weight="bold"),
                      command=self._clear).pack(side="right")
        ctk.CTkButton(header_inner, text="🔄 Lam moi", width=110, height=35, corner_radius=10,
                      fg_color=("#6366f1", "#4f46e5"), hover_color=("#818cf8", "#6366f1"),
                      command=self._refresh).pack(side="right", padx=10)
        
        # History Content Card
        content_card = ctk.CTkFrame(container, corner_radius=15)
        content_card.pack(fill="both", expand=True)
        
        self.history_text = ctk.CTkTextbox(content_card, font=ctk.CTkFont(size=12, family="Consolas"),
                                            corner_radius=10)
        self.history_text.pack(fill="both", expand=True, padx=15, pady=15)
        
        self._refresh()
    
    def _refresh(self):
        self.history_text.delete("1.0", "end")
        history = load_history()
        
        if not history:
            self.history_text.insert("1.0", "📭 Chua co lich su nao.\n\nLich su se duoc luu khi ban chay cac tac vu.")
            return
        
        for entry in reversed(history):
            status_icon = "✅" if entry["status"] == "SUCCESS" else "❌"
            type_badge = f"[{entry['type']}]".ljust(8)
            line = f"{entry['time']}  {status_icon} {type_badge} {entry['message']}\n"
            self.history_text.insert("end", line)
    
    def _clear(self):
        if messagebox.askyesno("Xac nhan", "Xoa toan bo lich su hoat dong?"):
            save_history([])
            self._refresh()
            messagebox.showinfo("Thanh cong", "Da xoa lich su!")

# =====================================================
# ============== CRON SETUP WINDOW ====================
# =====================================================
def get_script_path():
    """Get path to the main script"""
    if getattr(sys, 'frozen', False):
        # Running as compiled exe
        return sys.executable
    else:
        # Running as script
        return os.path.abspath(__file__)

def get_python_path():
    """Get Python executable path"""
    return sys.executable

def check_cron_installed():
    """Check if cronjob is already installed (Linux/macOS)"""
    if sys.platform == "win32":
        return None  # Windows uses Task Scheduler
    
    try:
        result = subprocess.run(['crontab', '-l'], capture_output=True, text=True)
        if result.returncode == 0:
            return 'uda_tools.py --hrm-auto' in result.stdout or \
                   get_script_path() in result.stdout
    except:
        pass
    return False

def setup_cron_linux(hour, minute, days=None):
    """Setup crontab on Linux/macOS with specific days"""
    script_path = get_script_path()
    python_path = get_python_path()
    
    # Convert days to cron format (0=Sun, 1=Mon, ... 6=Sat)
    if days and len(days) < 7:
        days_str = ",".join(str(d) for d in sorted(days))
    else:
        days_str = "*"
    
    # Build cron line
    if getattr(sys, 'frozen', False):
        cron_cmd = f"{minute} {hour} * * {days_str} {script_path} --hrm-auto"
    else:
        cron_cmd = f"{minute} {hour} * * {days_str} {python_path} {script_path} --hrm-auto"
    
    try:
        # Get existing crontab
        result = subprocess.run(['crontab', '-l'], capture_output=True, text=True, check=False)
        existing = result.stdout if result.returncode == 0 else ""
        
        # Remove old HRM entries
        lines = [l for l in existing.split('\n') if l.strip() and '--hrm-auto' not in l]
        lines.append(cron_cmd)
        
        # Write new crontab
        new_crontab = '\n'.join(lines) + '\n'
        proc = subprocess.Popen(['crontab', '-'], stdin=subprocess.PIPE, text=True)
        proc.communicate(input=new_crontab)
        
        # Format day names for message
        day_names = {0: "CN", 1: "T2", 2: "T3", 3: "T4", 4: "T5", 5: "T6", 6: "T7"}
        if days and len(days) < 7:
            days_display = ", ".join(day_names[d] for d in sorted(days))
        else:
            days_display = "Moi ngay"
        
        return True, f"Da cai cron: {hour:02d}:{minute:02d} ({days_display})"
    except Exception as e:
        return False, str(e)

def remove_cron_linux():
    """Remove cron entry on Linux/macOS"""
    try:
        result = subprocess.run(['crontab', '-l'], capture_output=True, text=True)
        if result.returncode != 0:
            return True, "Khong co cron nao"
        
        lines = [l for l in result.stdout.split('\n') if l.strip() and '--hrm-auto' not in l]
        
        if lines:
            new_crontab = '\n'.join(lines) + '\n'
            proc = subprocess.Popen(['crontab', '-'], stdin=subprocess.PIPE, text=True)
            proc.communicate(input=new_crontab)
        else:
            subprocess.run(['crontab', '-r'], capture_output=True)
        
        return True, "Da xoa cron thanh cong"
    except Exception as e:
        return False, str(e)

def setup_task_windows(hour, minute, days=None):
    """Setup Task Scheduler on Windows with specific days"""
    script_path = get_script_path()
    task_name = "UDA_HRM_AutoCheckin"
    
    # Map day numbers to Windows day names
    day_map = {0: "SUN", 1: "MON", 2: "TUE", 3: "WED", 4: "THU", 5: "FRI", 6: "SAT"}
    
    try:
        # Delete existing task
        subprocess.run(['schtasks', '/delete', '/tn', task_name, '/f'], 
                      capture_output=True, check=False)
        
        # Determine schedule type
        if days and len(days) < 7:
            days_str = ",".join(day_map[d] for d in sorted(days))
            schedule = f"/sc weekly /d {days_str}"
        else:
            schedule = "/sc daily"
        
        # Create new task
        if getattr(sys, 'frozen', False):
            cmd = f'schtasks /create /tn "{task_name}" /tr "\\"{script_path}\\" --hrm-auto" {schedule} /st {hour:02d}:{minute:02d} /f'
        else:
            python_path = get_python_path()
            cmd = f'schtasks /create /tn "{task_name}" /tr "\\"{python_path}\\" \\"{script_path}\\" --hrm-auto" {schedule} /st {hour:02d}:{minute:02d} /f'
        
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, check=False)
        
        if result.returncode == 0:
            day_names = {0: "CN", 1: "T2", 2: "T3", 3: "T4", 4: "T5", 5: "T6", 6: "T7"}
            if days and len(days) < 7:
                days_display = ", ".join(day_names[d] for d in sorted(days))
            else:
                days_display = "Moi ngay"
            return True, f"Da tao Task: {hour:02d}:{minute:02d} ({days_display})"
        else:
            return False, result.stderr
    except Exception as e:
        return False, str(e)

def remove_task_windows():
    """Remove Task Scheduler task on Windows"""
    task_name = "UDA_HRM_AutoCheckin"
    try:
        result = subprocess.run(['schtasks', '/delete', '/tn', task_name, '/f'], 
                               capture_output=True, text=True)
        if result.returncode == 0:
            return True, "Da xoa Task thanh cong"
        else:
            return True, "Khong co Task nao"
    except Exception as e:
        return False, str(e)

def setup_launchd_macos(hour, minute, days=None):
    """Setup launchd on macOS with specific days"""
    script_path = get_script_path()
    python_path = get_python_path()
    plist_path = os.path.expanduser("~/Library/LaunchAgents/com.uda.hrm.plist")
    
    if getattr(sys, 'frozen', False):
        program_args = f"""    <array>
        <string>{script_path}</string>
        <string>--hrm-auto</string>
    </array>"""
    else:
        program_args = f"""    <array>
        <string>{python_path}</string>
        <string>{script_path}</string>
        <string>--hrm-auto</string>
    </array>"""
    
    # Build StartCalendarInterval for selected days
    if days and len(days) < 7:
        intervals = "\n".join([f"""        <dict>
            <key>Hour</key>
            <integer>{hour}</integer>
            <key>Minute</key>
            <integer>{minute}</integer>
            <key>Weekday</key>
            <integer>{d}</integer>
        </dict>""" for d in sorted(days)])
        calendar_interval = f"""    <array>
{intervals}
    </array>"""
    else:
        calendar_interval = f"""    <dict>
        <key>Hour</key>
        <integer>{hour}</integer>
        <key>Minute</key>
        <integer>{minute}</integer>
    </dict>"""
    
    plist_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN"
"http://www.apple.com/DTDs/PropertyList-1.0.dtd">
<plist version="1.0">
<dict>
    <key>Label</key>
    <string>com.uda.hrm</string>
    <key>ProgramArguments</key>
{program_args}
    <key>StartCalendarInterval</key>
{calendar_interval}
    <key>StandardOutPath</key>
    <string>{APP_DIR}/launchd.log</string>
    <key>StandardErrorPath</key>
    <string>{APP_DIR}/launchd_error.log</string>
</dict>
</plist>
"""
    
    try:
        # Unload existing
        subprocess.run(['launchctl', 'unload', plist_path], capture_output=True, check=False)
        
        # Write plist
        os.makedirs(os.path.dirname(plist_path), exist_ok=True)
        with open(plist_path, 'w', encoding='utf-8') as f:
            f.write(plist_content)
        
        # Load new
        result = subprocess.run(['launchctl', 'load', plist_path], capture_output=True, text=True, check=False)
        
        if result.returncode == 0:
            day_names = {0: "CN", 1: "T2", 2: "T3", 3: "T4", 4: "T5", 5: "T6", 6: "T7"}
            if days and len(days) < 7:
                days_display = ", ".join(day_names[d] for d in sorted(days))
            else:
                days_display = "Moi ngay"
            return True, f"Da cai launchd: {hour:02d}:{minute:02d} ({days_display})"
        else:
            return False, result.stderr
    except Exception as e:
        return False, str(e)

def remove_launchd_macos():
    """Remove launchd on macOS"""
    plist_path = os.path.expanduser("~/Library/LaunchAgents/com.uda.hrm.plist")
    try:
        subprocess.run(['launchctl', 'unload', plist_path], capture_output=True)
        if os.path.exists(plist_path):
            os.remove(plist_path)
        return True, "Da xoa launchd thanh cong"
    except Exception as e:
        return False, str(e)

class CronSetupWindow(ctk.CTkToplevel):
    # Day names for display
    DAYS = [
        ("T2", 1), ("T3", 2), ("T4", 3), ("T5", 4), 
        ("T6", 5), ("T7", 6), ("CN", 0)
    ]
    
    def __init__(self, parent, config, on_save):
        super().__init__(parent)
        self.parent_app = parent
        self.config = config
        self.on_save = on_save
        
        self.title("Cai dat lich tu dong")
        self.geometry("550x520")
        self.resizable(False, False)
        
        # Day selection vars
        self.day_vars = {}
        
        self._build_ui()
        
        self.transient(parent)
        self.update()
        x = parent.winfo_x() + (parent.winfo_width() - 550) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 520) // 2
        self.geometry(f"550x520+{x}+{y}")
        self.lift()
        self.focus_force()
    
    def _build_ui(self):
        # OS Info badge (không cần header vì đã có title bar)
        if sys.platform == "win32":
            os_name = "Windows Task Scheduler"
            os_icon = "🪟"
        elif sys.platform == "darwin":
            os_name = "macOS launchd"
            os_icon = "🍎"
        else:
            os_name = "Linux crontab"
            os_icon = "🐧"
        
        os_badge = ctk.CTkFrame(self, fg_color=("#e0e7ff", "#312e81"), corner_radius=20)
        os_badge.pack(pady=(20, 10))
        ctk.CTkLabel(os_badge, text=f" {os_icon} {os_name} ", 
                     font=ctk.CTkFont(size=12)).pack(padx=15, pady=5)
        
        # Main content
        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Time Selection Card
        time_card = ctk.CTkFrame(content, corner_radius=15)
        time_card.pack(fill="x", pady=10)
        
        ctk.CTkLabel(time_card, text="🕐 Thoi gian chay", 
                     font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=20, pady=(15, 10))
        
        time_row = ctk.CTkFrame(time_card, fg_color="transparent")
        time_row.pack(fill="x", padx=20, pady=(0, 15))
        
        # Hour picker
        hour_frame = ctk.CTkFrame(time_row, fg_color=("gray90", "gray25"), corner_radius=10)
        hour_frame.pack(side="left", padx=5)
        ctk.CTkLabel(hour_frame, text="Gio", font=ctk.CTkFont(size=11), 
                     text_color="gray").pack(padx=10, pady=(5, 0))
        self.hour_var = ctk.StringVar(value="8")
        ctk.CTkComboBox(hour_frame, values=[f"{i:02d}" for i in range(24)],
                        variable=self.hour_var, width=80, 
                        font=ctk.CTkFont(size=16, weight="bold")).pack(padx=10, pady=(0, 8))
        
        ctk.CTkLabel(time_row, text=":", font=ctk.CTkFont(size=24, weight="bold")).pack(side="left", padx=5)
        
        # Minute picker
        minute_frame = ctk.CTkFrame(time_row, fg_color=("gray90", "gray25"), corner_radius=10)
        minute_frame.pack(side="left", padx=5)
        ctk.CTkLabel(minute_frame, text="Phut", font=ctk.CTkFont(size=11), 
                     text_color="gray").pack(padx=10, pady=(5, 0))
        self.minute_var = ctk.StringVar(value="00")
        ctk.CTkComboBox(minute_frame, values=[f"{i:02d}" for i in range(0, 60, 5)],
                        variable=self.minute_var, width=80,
                        font=ctk.CTkFont(size=16, weight="bold")).pack(padx=10, pady=(0, 8))
        
        # Days Selection Card
        days_card = ctk.CTkFrame(content, corner_radius=15)
        days_card.pack(fill="x", pady=10)
        
        ctk.CTkLabel(days_card, text="📅 Ngay trong tuan", 
                     font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=20, pady=(15, 10))
        
        days_row = ctk.CTkFrame(days_card, fg_color="transparent")
        days_row.pack(fill="x", padx=15, pady=(0, 15))
        
        for name, day_num in self.DAYS:
            var = ctk.BooleanVar(value=day_num in [1, 2, 3, 4, 5])  # Default: Mon-Fri
            self.day_vars[day_num] = var
            
            # Custom styled checkbox
            day_btn = ctk.CTkCheckBox(days_row, text=name, variable=var, width=50,
                                       checkbox_width=20, checkbox_height=20,
                                       font=ctk.CTkFont(size=12, weight="bold"))
            day_btn.pack(side="left", padx=5)
        
        # Quick select buttons
        quick_row = ctk.CTkFrame(days_card, fg_color="transparent")
        quick_row.pack(fill="x", padx=20, pady=(0, 15))
        
        ctk.CTkButton(quick_row, text="T2-T6", width=70, height=28,
                      fg_color=("gray70", "gray40"), hover_color=("gray60", "gray50"),
                      font=ctk.CTkFont(size=11),
                      command=self._select_weekdays).pack(side="left", padx=3)
        ctk.CTkButton(quick_row, text="Tat ca", width=70, height=28,
                      fg_color=("gray70", "gray40"), hover_color=("gray60", "gray50"),
                      font=ctk.CTkFont(size=11),
                      command=self._select_all_days).pack(side="left", padx=3)
        ctk.CTkButton(quick_row, text="Bo chon", width=70, height=28,
                      fg_color=("gray70", "gray40"), hover_color=("gray60", "gray50"),
                      font=ctk.CTkFont(size=11),
                      command=self._clear_days).pack(side="left", padx=3)
        
        # Status Card
        status_card = ctk.CTkFrame(content, corner_radius=15, fg_color=("gray95", "gray20"))
        status_card.pack(fill="x", pady=10)
        
        self.status_label = ctk.CTkLabel(status_card, text="⚪ Chua cai dat", 
                                          font=ctk.CTkFont(size=13),
                                          text_color="gray")
        self.status_label.pack(pady=12)
        
        self._check_status()
        
        # Action Buttons
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x", pady=10)
        
        ctk.CTkButton(btn_frame, text="✅ CAI DAT", width=150, height=45,
                      fg_color=("#22c55e", "#16a34a"), hover_color=("#15803d", "#14532d"),
                      font=ctk.CTkFont(size=14, weight="bold"),
                      command=self._install).pack(side="left", padx=5)
        
        ctk.CTkButton(btn_frame, text="🗑️ GO BO", width=120, height=45,
                      fg_color=("#ef4444", "#dc2626"), hover_color=("#b91c1c", "#991b1b"),
                      font=ctk.CTkFont(size=14, weight="bold"),
                      command=self._uninstall).pack(side="left", padx=5)
        
        ctk.CTkButton(btn_frame, text="Dong", width=100, height=45,
                      fg_color=("gray60", "gray40"),
                      font=ctk.CTkFont(size=14),
                      command=self.destroy).pack(side="right", padx=5)
        
        # Warning note
        note_frame = ctk.CTkFrame(content, fg_color=("#fef3c7", "#451a03"), corner_radius=10)
        note_frame.pack(fill="x", pady=5)
        ctk.CTkLabel(note_frame, text="⚠️ Nho tick 'Nho mat khau' va chay 1 lan truoc khi cai dat!", 
                     font=ctk.CTkFont(size=11),
                     text_color=("#92400e", "#fbbf24")).pack(pady=8)
    
    def _select_weekdays(self):
        for day_num, var in self.day_vars.items():
            var.set(day_num in [1, 2, 3, 4, 5])
    
    def _select_all_days(self):
        for var in self.day_vars.values():
            var.set(True)
    
    def _clear_days(self):
        for var in self.day_vars.values():
            var.set(False)
    
    def _get_selected_days(self):
        """Return list of selected day numbers (0=Sun, 1=Mon, ... 6=Sat)"""
        return [day for day, var in self.day_vars.items() if var.get()]
    
    def _days_to_cron(self, days):
        """Convert day numbers to cron format"""
        if not days:
            return "*"
        if len(days) == 7:
            return "*"
        return ",".join(str(d) for d in sorted(days))
    
    def _check_status(self):
        if sys.platform == "win32":
            try:
                result = subprocess.run(['schtasks', '/query', '/tn', 'UDA_HRM_AutoCheckin'],
                                       capture_output=True, text=True, check=False)
                if result.returncode == 0:
                    self.status_label.configure(text="🟢 Da cai dat Task Scheduler", text_color="#22c55e")
                    return
            except Exception:
                pass
        elif sys.platform == "darwin":
            plist_path = os.path.expanduser("~/Library/LaunchAgents/com.uda.hrm.plist")
            if os.path.exists(plist_path):
                self.status_label.configure(text="🟢 Da cai dat launchd", text_color="#22c55e")
                return
        else:
            if check_cron_installed():
                self.status_label.configure(text="🟢 Da cai dat crontab", text_color="#22c55e")
                return
        
        self.status_label.configure(text="⚪ Chua cai dat", text_color="gray")
    
    def _install(self):
        if not self.config.get("hrm_password"):
            messagebox.showwarning("Thieu mat khau", 
                "Ban chua luu mat khau!\n\n"
                "1. Quay lai tab HRM\n"
                "2. Tick 'Nho mat khau (Cronjob)'\n"
                "3. Click 'CHAY NGAY' it nhat 1 lan\n"
                "4. Quay lai day cai dat")
            return
        
        days = self._get_selected_days()
        if not days:
            messagebox.showwarning("Chua chon ngay", "Vui long chon it nhat 1 ngay!")
            return
        
        hour = int(self.hour_var.get())
        minute = int(self.minute_var.get())
        
        if sys.platform == "win32":
            success, msg = setup_task_windows(hour, minute, days)
        elif sys.platform == "darwin":
            success, msg = setup_launchd_macos(hour, minute, days)
        else:
            success, msg = setup_cron_linux(hour, minute, days)
        
        if success:
            messagebox.showinfo("Thanh cong", msg)
            self._check_status()
        else:
            messagebox.showerror("Loi", f"Khong the cai dat:\n{msg}")
    
    def _uninstall(self):
        if not messagebox.askyesno("Xac nhan", "Go bo cronjob/task da cai?"):
            return
        
        if sys.platform == "win32":
            success, msg = remove_task_windows()
        elif sys.platform == "darwin":
            success, msg = remove_launchd_macos()
        else:
            success, msg = remove_cron_linux()
        
        if success:
            messagebox.showinfo("Thanh cong", msg)
            self._check_status()
        else:
            messagebox.showerror("Loi", f"Khong the go bo:\n{msg}")

# =====================================================
# ============== MAIN APP =============================
# =====================================================
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(f"UDA Tools Pro v{VERSION}")
        self.geometry("950x800")
        self.minsize(800, 600)  # Minimum size
        self.resizable(True, True)  # Allow fullscreen/resize
        
        self.config = load_config()
        
        self._build_ui()
    
    def _bind_mousewheel(self, scrollable_frame):
        """Bind mouse wheel scrolling for all platforms"""
        canvas = scrollable_frame._parent_canvas
        
        def _scroll_linux_up(event):
            canvas.yview_scroll(-1, "units")
            return "break"
        
        def _scroll_linux_down(event):
            canvas.yview_scroll(1, "units")
            return "break"
        
        def _scroll_windows(event):
            canvas.yview_scroll(int(-event.delta/60), "units")  # Smoother scroll
            return "break"
        
        def _on_enter(event):
            # Linux
            scrollable_frame.bind_all("<Button-4>", _scroll_linux_up)
            scrollable_frame.bind_all("<Button-5>", _scroll_linux_down)
            # Windows/macOS  
            scrollable_frame.bind_all("<MouseWheel>", _scroll_windows)
        
        def _on_leave(event):
            scrollable_frame.unbind_all("<Button-4>")
            scrollable_frame.unbind_all("<Button-5>")
            scrollable_frame.unbind_all("<MouseWheel>")
        
        # Bind to the scrollable frame itself
        scrollable_frame.bind("<Enter>", _on_enter)
        scrollable_frame.bind("<Leave>", _on_leave)
        
        # Also bind to all children
        def bind_children(widget):
            widget.bind("<Enter>", _on_enter)
            for child in widget.winfo_children():
                bind_children(child)
        
        # Schedule binding after widgets are created
        scrollable_frame.after(100, lambda: bind_children(scrollable_frame))
    
    def _build_ui(self):
        # Gradient Header
        header = ctk.CTkFrame(self, corner_radius=0, fg_color=("#4f46e5", "#3730a3"), height=70)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        header_inner = ctk.CTkFrame(header, fg_color="transparent")
        header_inner.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Logo + Title
        title_frame = ctk.CTkFrame(header_inner, fg_color="transparent")
        title_frame.pack(side="left", fill="y")
        
        ctk.CTkLabel(title_frame, text="🚀 UDA TOOLS PRO", 
                     font=ctk.CTkFont(size=26, weight="bold"),
                     text_color="white").pack(side="left")
        ctk.CTkLabel(title_frame, text=f"v{VERSION}", 
                     font=ctk.CTkFont(size=12),
                     text_color="#c7d2fe").pack(side="left", padx=10, pady=5)
        
        # Header buttons
        btn_frame = ctk.CTkFrame(header_inner, fg_color="transparent")
        btn_frame.pack(side="right")
        
        ctk.CTkButton(btn_frame, text="⚙️ Cai dat", width=100, height=35,
                      fg_color=("#818cf8", "#6366f1"), hover_color=("#a5b4fc", "#818cf8"),
                      font=ctk.CTkFont(size=13),
                      command=lambda: SettingsWindow(self, self.config, self._on_settings_save)
                      ).pack(side="right", padx=5)
        ctk.CTkButton(btn_frame, text="📋 Lich su", width=100, height=35,
                      fg_color=("#818cf8", "#6366f1"), hover_color=("#a5b4fc", "#818cf8"),
                      font=ctk.CTkFont(size=13),
                      command=lambda: HistoryWindow(self)).pack(side="right", padx=5)
        
        # Tabview with custom styling
        self.tabview = ctk.CTkTabview(self, corner_radius=15,
                                       fg_color=("gray95", "gray15"),
                                       segmented_button_fg_color=("#e0e7ff", "#312e81"),
                                       segmented_button_selected_color=("#6366f1", "#4f46e5"),
                                       segmented_button_selected_hover_color=("#818cf8", "#6366f1"))
        self.tabview.pack(fill="both", expand=True, padx=15, pady=15)
        
        self.tab_uda = self.tabview.add("📊 Nhap Diem UDA")
        self.tab_hrm = self.tabview.add("⏰ HRM Auto")
        
        self._build_uda_tab()
        self._build_hrm_tab()
    
    def _on_settings_save(self, new_config):
        self.config = new_config
        save_config(self.config)
    
    # ==================== UDA TAB ====================
    def _build_uda_tab(self):
        # Use scrollable frame for responsive
        tab = ctk.CTkScrollableFrame(self.tab_uda, fg_color="transparent")
        tab.pack(fill="both", expand=True)
        tab.grid_columnconfigure(0, weight=1)
        self._bind_mousewheel(tab)  # Enable mouse wheel scrolling
        
        # Login Card
        login_card = ctk.CTkFrame(tab, corner_radius=15)
        login_card.pack(fill="x", padx=10, pady=10)
        login_card.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(login_card, text="🔐 THONG TIN DANG NHAP", 
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=("#4f46e5", "#a5b4fc")).grid(row=0, column=0, columnspan=2, 
                                                            sticky="w", padx=20, pady=(15, 10))
        
        # Username
        ctk.CTkLabel(login_card, text="👤 Tai khoan:", font=ctk.CTkFont(size=12)).grid(row=1, column=0, sticky="w", padx=20, pady=8)
        self.uda_user = ctk.CTkEntry(login_card, placeholder_text="Nhap tai khoan", height=38, corner_radius=10)
        self.uda_user.grid(row=1, column=1, sticky="ew", padx=20, pady=8)
        if self.config.get("uda_username"):
            self.uda_user.insert(0, self.config["uda_username"])
        
        # Password
        ctk.CTkLabel(login_card, text="🔑 Mat khau:", font=ctk.CTkFont(size=12)).grid(row=2, column=0, sticky="w", padx=20, pady=8)
        self.uda_pass = ctk.CTkEntry(login_card, placeholder_text="Nhap mat khau", show="*", height=38, corner_radius=10)
        self.uda_pass.grid(row=2, column=1, sticky="ew", padx=20, pady=8)
        
        # Subject
        ctk.CTkLabel(login_card, text="📚 Ma mon:", font=ctk.CTkFont(size=12)).grid(row=3, column=0, sticky="w", padx=20, pady=8)
        self.uda_subject = ctk.CTkEntry(login_card, placeholder_text="VD: Ky nang so (1tc)/OK//93190/7481/KL24A", height=38, corner_radius=10)
        self.uda_subject.grid(row=3, column=1, sticky="ew", padx=20, pady=8)
        if self.config.get("uda_subject"):
            self.uda_subject.insert(0, self.config["uda_subject"])
        
        # File
        ctk.CTkLabel(login_card, text="📄 File diem:", font=ctk.CTkFont(size=12)).grid(row=4, column=0, sticky="w", padx=20, pady=8)
        file_frame = ctk.CTkFrame(login_card, fg_color="transparent")
        file_frame.grid(row=4, column=1, sticky="ew", padx=20, pady=8)
        file_frame.grid_columnconfigure(0, weight=1)
        
        self.uda_file = ctk.CTkEntry(file_frame, placeholder_text="Chon file Excel...", state="disabled", height=38, corner_radius=10)
        self.uda_file.grid(row=0, column=0, sticky="ew")
        ctk.CTkButton(file_frame, text="📂 Chon", width=80, height=38, corner_radius=10,
                      fg_color=("#6366f1", "#4f46e5"),
                      command=self._browse_excel).grid(row=0, column=1, padx=(8,0))
        ctk.CTkButton(file_frame, text="📥 Tai mau", width=90, height=38, corner_radius=10,
                      fg_color=("#f59e0b", "#d97706"), hover_color=("#d97706", "#b45309"),
                      command=self._download_template).grid(row=0, column=2, padx=(8,0))
        
        # Options Card
        options_card = ctk.CTkFrame(tab, corner_radius=15)
        options_card.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(options_card, text="⚙️ TUY CHON", 
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=("#4f46e5", "#a5b4fc")).pack(anchor="w", padx=20, pady=(15, 10))
        
        opts_inner = ctk.CTkFrame(options_card, fg_color="transparent")
        opts_inner.pack(fill="x", padx=20, pady=(0, 15))
        
        self.uda_headless = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(opts_inner, text="🔇 Chay an (Headless)", 
                        variable=self.uda_headless, corner_radius=5).pack(side="left", padx=10)
        
        self.uda_save_config = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(opts_inner, text="💾 Nho thong tin", 
                        variable=self.uda_save_config, corner_radius=5).pack(side="left", padx=10)
        
        # Columns Card
        cols_card = ctk.CTkFrame(tab, corner_radius=15)
        cols_card.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(cols_card, text="📊 COT DIEM CAN NHAP", 
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=("#4f46e5", "#a5b4fc")).pack(anchor="w", padx=20, pady=(15, 10))
        
        cols_inner = ctk.CTkFrame(cols_card, fg_color="transparent")
        cols_inner.pack(fill="x", padx=20, pady=(0, 15))
        
        self.uda_cols = {}
        for idx, title in enumerate(ALL_TITLES):
            var = ctk.BooleanVar(value=title in DEFAULT_TITLES)
            chk = ctk.CTkCheckBox(cols_inner, text=title, variable=var, 
                                   checkbox_width=22, checkbox_height=22,
                                   font=ctk.CTkFont(size=12, weight="bold"))
            chk.grid(row=0, column=idx, padx=8, pady=5)
            self.uda_cols[title] = var
        
        # Progress
        progress_frame = ctk.CTkFrame(tab, fg_color="transparent")
        progress_frame.pack(fill="x", padx=20, pady=10)
        
        self.uda_progress = ctk.CTkProgressBar(progress_frame, height=12, corner_radius=6,
                                                progress_color=("#22c55e", "#16a34a"))
        self.uda_progress.pack(fill="x")
        self.uda_progress.set(0)
        
        # Buttons
        btn_frame = ctk.CTkFrame(tab, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=10)
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        
        self.btn_uda_import = ctk.CTkButton(btn_frame, text="✅ NHAP DIEM", height=50,
                                            font=ctk.CTkFont(size=15, weight="bold"),
                                            fg_color=("#22c55e", "#16a34a"),
                                            hover_color=("#15803d", "#14532d"),
                                            corner_radius=12,
                                            command=lambda: self._run_uda(False))
        self.btn_uda_import.grid(row=0, column=0, sticky="ew", padx=5)
        
        self.btn_uda_delete = ctk.CTkButton(btn_frame, text="🗑️ XOA DIEM", height=50,
                                            font=ctk.CTkFont(size=15, weight="bold"),
                                            fg_color=("#ef4444", "#dc2626"),
                                            hover_color=("#b91c1c", "#991b1b"),
                                            corner_radius=12,
                                            command=lambda: self._run_uda(True))
        self.btn_uda_delete.grid(row=0, column=1, sticky="ew", padx=5)
        
        # Status
        self.uda_status = ctk.CTkLabel(tab, text="✨ San sang...", 
                                        font=ctk.CTkFont(size=12),
                                        text_color=("#6b7280", "#9ca3af"))
        self.uda_status.pack(pady=5)
    
    def _browse_excel(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if f:
            self.uda_file.configure(state="normal")
            self.uda_file.delete(0, "end")
            self.uda_file.insert(0, f)
            self.uda_file.configure(state="disabled")
    
    def _download_template(self):
        """Create and save a template Excel file"""
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfilename="mau_nhap_diem.xlsx"
        )
        if not save_path:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Nhap Diem"
            
            # Headers
            headers = ["MSSV", "KTTX", "CCAN", "GHP", "TDNH", "THTN", "TLDA", "THI1"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Sample data
            sample_data = [
                ["SV001", 8, 7, 9, "", "", "", 8],
                ["SV002", 7, 8, 8, "", "", "", 7],
            ]
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(save_path)
            messagebox.showinfo("Thanh cong", f"Da tao file mau:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Loi", f"Khong the tao file mau:\n{str(e)}")
    
    def _run_uda(self, is_delete):
        if is_delete and not messagebox.askyesno("Xac nhan", "Xoa diem?"):
            return
        
        username = self.uda_user.get().strip()
        password = self.uda_pass.get()
        subject = self.uda_subject.get().strip()
        filepath = self.uda_file.get()
        
        if not all([username, password, subject, filepath]):
            messagebox.showwarning("Thieu thong tin", "Vui long nhap day du!")
            return
        
        selected = [t for t, v in self.uda_cols.items() if v.get()]
        if not selected:
            messagebox.showwarning("Loi", "Chua chon cot diem!")
            return
        
        if self.uda_save_config.get():
            self.config["uda_username"] = username
            self.config["uda_subject"] = subject
            save_config(self.config)
        
        self.btn_uda_import.configure(state="disabled")
        self.btn_uda_delete.configure(state="disabled")
        
        def run():
            run_uda_grader(
                username, password, subject, filepath, selected,
                self._update_uda_status, self._update_uda_progress,
                is_delete, self.uda_headless.get(),
                self.config.get("chrome_path")
            )
            self.btn_uda_import.configure(state="normal")
            self.btn_uda_delete.configure(state="normal")
            self.uda_progress.set(0)
        
        threading.Thread(target=run, daemon=True).start()
    
    def _update_uda_status(self, text):
        self.uda_status.configure(text=text)
        self.update_idletasks()
    
    def _update_uda_progress(self, value):
        self.uda_progress.set(value / 100)
        self.update_idletasks()
    
    # ==================== HRM TAB ====================
    def _build_hrm_tab(self):
        # Use scrollable frame for responsive
        tab = ctk.CTkScrollableFrame(self.tab_hrm, fg_color="transparent")
        tab.pack(fill="both", expand=True)
        self._bind_mousewheel(tab)  # Enable mouse wheel scrolling
        
        # Login Card
        login_card = ctk.CTkFrame(tab, corner_radius=15)
        login_card.pack(fill="x", padx=10, pady=10)
        login_card.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(login_card, text="🔐 THONG TIN DANG NHAP HRM", 
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=("#4f46e5", "#a5b4fc")).grid(row=0, column=0, columnspan=2, 
                                                            sticky="w", padx=20, pady=(15, 10))
        
        # Email
        ctk.CTkLabel(login_card, text="📧 Email:", font=ctk.CTkFont(size=12)).grid(row=1, column=0, sticky="w", padx=20, pady=8)
        self.hrm_user = ctk.CTkEntry(login_card, placeholder_text="email@donga.edu.vn", height=38, corner_radius=10)
        self.hrm_user.grid(row=1, column=1, sticky="ew", padx=20, pady=8)
        if self.config.get("hrm_username"):
            self.hrm_user.insert(0, self.config["hrm_username"])
        
        # Password
        ctk.CTkLabel(login_card, text="🔑 Mat khau:", font=ctk.CTkFont(size=12)).grid(row=2, column=0, sticky="w", padx=20, pady=8)
        self.hrm_pass = ctk.CTkEntry(login_card, placeholder_text="Nhap mat khau", show="*", height=38, corner_radius=10)
        self.hrm_pass.grid(row=2, column=1, sticky="ew", padx=20, pady=(8, 15))
        
        # Content Card
        content_card = ctk.CTkFrame(tab, corner_radius=15)
        content_card.pack(fill="both", expand=True, padx=10, pady=5)
        
        ctk.CTkLabel(content_card, text="📝 NOI DUNG CONG VIEC", 
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=("#4f46e5", "#a5b4fc")).pack(anchor="w", padx=20, pady=(15, 5))
        ctk.CTkLabel(content_card, text="(Moi dong = 1 noi dung, he thong se random khi chay)", 
                     font=ctk.CTkFont(size=11),
                     text_color=("#6b7280", "#9ca3af")).pack(anchor="w", padx=20, pady=(0, 10))
        
        self.hrm_content = ctk.CTkTextbox(content_card, height=150, corner_radius=10,
                                           font=ctk.CTkFont(size=12))
        self.hrm_content.pack(fill="both", expand=True, padx=20, pady=(0, 15))
        
        # Load saved content
        default_content = self.config.get("hrm_contents", 
            "Soan noi dung thuc hanh\nHo tro sinh vien\nCham bai tap\nSoan de thi")
        self.hrm_content.insert("1.0", default_content)
        
        # Options Card
        options_card = ctk.CTkFrame(tab, corner_radius=15)
        options_card.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(options_card, text="⚙️ TUY CHON & LICH TU DONG", 
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=("#4f46e5", "#a5b4fc")).pack(anchor="w", padx=20, pady=(15, 10))
        
        opts_inner = ctk.CTkFrame(options_card, fg_color="transparent")
        opts_inner.pack(fill="x", padx=20, pady=(0, 15))
        
        self.hrm_headless = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(opts_inner, text="🔇 Chay an", 
                        variable=self.hrm_headless, corner_radius=5).pack(side="left", padx=8)
        
        self.hrm_save_config = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(opts_inner, text="💾 Nho thong tin", 
                        variable=self.hrm_save_config, corner_radius=5).pack(side="left", padx=8)
        
        self.hrm_save_pass = ctk.BooleanVar(value=bool(self.config.get("hrm_password")))
        ctk.CTkCheckBox(opts_inner, text="🔑 Nho mat khau", 
                        variable=self.hrm_save_pass, corner_radius=5,
                        text_color=("#f59e0b", "#fbbf24")).pack(side="left", padx=8)
        
        ctk.CTkButton(opts_inner, text="📅 Cai Cronjob", width=130, height=35,
                      fg_color=("#8b5cf6", "#7c3aed"), hover_color=("#7c3aed", "#6d28d9"),
                      corner_radius=10, font=ctk.CTkFont(weight="bold"),
                      command=self._open_cron_setup).pack(side="right", padx=5)
        
        # Action Button
        btn_frame = ctk.CTkFrame(tab, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=10)
        
        self.btn_hrm_run = ctk.CTkButton(btn_frame, text="🚀 CHAY NGAY", height=50,
                                         font=ctk.CTkFont(size=16, weight="bold"),
                                         fg_color=("#22c55e", "#16a34a"),
                                         hover_color=("#15803d", "#14532d"),
                                         corner_radius=12,
                                         command=self._run_hrm)
        self.btn_hrm_run.pack(fill="x", padx=5)
        
        # Status
        self.hrm_status = ctk.CTkLabel(tab, text="✨ San sang...", 
                                        font=ctk.CTkFont(size=12),
                                        text_color=("#6b7280", "#9ca3af"))
        self.hrm_status.pack(pady=5)
    
    def _open_cron_setup(self):
        CronSetupWindow(self, self.config, self._on_settings_save)
    
    def _run_hrm(self):
        username = self.hrm_user.get().strip()
        password = self.hrm_pass.get()
        contents = self.hrm_content.get("1.0", "end").strip().split("\n")
        contents = [c.strip() for c in contents if c.strip()]
        
        if not all([username, password]) or not contents:
            messagebox.showwarning("Thieu thong tin", "Vui long nhap day du!")
            return
        
        # Random select content
        task_title = random.choice(contents)
        task_detail = f"{task_title} - {datetime.now().strftime('%d/%m/%Y')}"
        
        # Save config
        if self.hrm_save_config.get():
            self.config["hrm_username"] = username
            self.config["hrm_contents"] = "\n".join(contents)
        
        # Save password for cronjob (optional)
        if self.hrm_save_pass.get():
            self.config["hrm_password"] = password
        else:
            self.config.pop("hrm_password", None)
        
        save_config(self.config)
        
        self.btn_hrm_run.configure(state="disabled")
        
        def run():
            run_hrm_checkin(
                username, password, task_title, task_detail,
                self._update_hrm_status,
                self.config.get("chrome_path"),
                self.hrm_headless.get()
            )
            self.btn_hrm_run.configure(state="normal")
        
        threading.Thread(target=run, daemon=True).start()
    
    def _update_hrm_status(self, text):
        self.hrm_status.configure(text=text)
        self.update_idletasks()

# =====================================================
# ============== CLI MODE =============================
# =====================================================
def run_hrm_auto():
    """Run HRM auto from CLI/Cronjob"""
    config = load_config()
    
    username = config.get("hrm_username")
    password = config.get("hrm_password")  # Need to save password for auto mode
    contents = config.get("hrm_contents", "").split("\n")
    contents = [c.strip() for c in contents if c.strip()]
    
    if not all([username, password]) or not contents:
        logger.error("HRM Auto: Missing config. Run GUI first to configure.")
        return
    
    task_title = random.choice(contents)
    task_detail = f"{task_title} - {datetime.now().strftime('%d/%m/%Y')}"
    
    success = run_hrm_checkin(
        username, password, task_title, task_detail,
        lambda x: logger.info(x),
        config.get("chrome_path"),
        headless=True
    )
    
    if success:
        logger.info("HRM Auto: Completed successfully!")
    else:
        logger.error("HRM Auto: Failed!")

# =====================================================
# ============== MAIN =================================
# =====================================================
if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--hrm-auto":
        run_hrm_auto()
    else:
        app = App()
        app.mainloop()
